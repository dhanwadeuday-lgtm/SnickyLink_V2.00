#!/usr/bin/env python3
"""
SNICKYLINK — Excel -> SQL importer for the 365-mission SNICKS_updated.xlsx dataset.
Implements the verification classifier + frequency mapping documented in
docs/habit-algorithm.md. Generates migrations/0003_snicks_data.sql (idempotent:
wipes existing snicks rows tied to old seed IDs, then inserts all 365 real missions).

Run: python3 scripts/import_snicks.py
"""
import openpyxl
import re
import sys

SRC = '/home/user/uploaded_files/SNICKS_updated.xlsx'
OUT = '/home/user/webapp/migrations/0003_snicks_data.sql'

FREQ_MAP = {
    'Daily': 'DAILY',
    'Weekly': 'WEEKLY',
    'Monthly': 'MONTHLY',
    'Challenge': 'CHALLENGE',
}

DIFF_MAP = {
    'Easy': 'EASY', 'Low': 'EASY', 'Low/Med': 'EASY',
    'Medium': 'MEDIUM', 'Med': 'MEDIUM',
    'High': 'HARD', 'Very High': 'HARD',
}

SAFETY_MAP = {
    'OK — standard safety review': 'OK',
    'REVIEW / CONSENT + SAFETY': 'REVIEW_CONSENT_SAFETY',
    'CONSENT / SKIP OPTION': 'CONSENT_SKIP_OPTION',
}

CATEGORY_BY_DOMINANT_PILLAR = {
    'communication': 'cat_communication',
    'emotional': 'cat_emotional',
    'efforts': 'cat_efforts',
    'trust': 'cat_trust',
}


def classify_verification(raw: str) -> str:
    """Deterministic keyword classifier -> one of the 4 functional verification_type values.
    Priority order documented in docs/habit-algorithm.md section 1.2."""
    r = raw.lower()
    if 'partner confirmation' in r:
        return 'PARTNER_CONFIRMATION'
    if any(k in r for k in ('mutual', 'simultaneous', 'independent answers', 'answer reveal', 'answer submission')):
        return 'MUTUAL_COMPLETION'
    if any(k in r for k in ('photo', 'video', 'voice', 'image')):
        return 'OPTIONAL_NON_SENSITIVE_EVIDENCE'
    return 'SELF_CONFIRMATION'


def sql_str(v):
    if v is None:
        return 'NULL'
    s = str(v).replace("'", "''")
    return f"'{s}'"


def duration_for(freq: str, difficulty: str) -> int:
    base = {'DAILY': 8, 'WEEKLY': 30, 'MONTHLY': 60, 'CHALLENGE': 15}[freq]
    bump = {'EASY': 0, 'MEDIUM': 10, 'HARD': 25}.get(difficulty, 0)
    return base + bump


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb['Final Master']
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    rows_by_freq = {'DAILY': [], 'WEEKLY': [], 'MONTHLY': [], 'CHALLENGE': []}
    unmapped_verification = set()

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if row[0] is None:
            continue
        mission_id = row[idx['Mission ID']]
        freq_raw = row[idx['Frequency']]
        freq = FREQ_MAP[freq_raw]
        title = row[idx['Title']]
        desc = row[idx['Description']]
        diff_raw = row[idx['Source Difficulty']]
        difficulty = DIFF_MAP.get(diff_raw, 'MEDIUM')
        xp = int(row[idx['Points / XP']])
        verif_raw = row[idx['Verification']]
        verif_type = classify_verification(verif_raw)
        verif_level = 'MEDIUM' if (row[idx['Verification Level']] or '').strip().lower() == 'medium' else 'LOW'
        safety_raw = row[idx['Safety Status']]
        safety_status = SAFETY_MAP.get(safety_raw, 'OK')
        privacy_rule = row[idx['Privacy Rule']]
        comm_pct = int(row[idx['Communication %']] or 0)
        emo_pct = int(row[idx['Emotional Connection %']] or 0)
        eff_pct = int(row[idx['Efforts %']] or 0)
        trust_pct = int(row[idx['Trust %']] or 0)
        upgrade = (row[idx['Recommended Verification Upgrade']] or '').strip()

        # 1.3 Upgrade override: force PARTNER_CONFIRMATION when spreadsheet recommends
        # in-app-message-box escalation and current classification is still self-only.
        if upgrade.startswith('In-app message box') and verif_type in ('SELF_CONFIRMATION', 'OPTIONAL_NON_SENSITIVE_EVIDENCE'):
            verif_type = 'PARTNER_CONFIRMATION'

        # dominant pillar -> category (ties broken by fixed priority comm>emo>eff>trust)
        pillars = [('communication', comm_pct), ('emotional', emo_pct), ('efforts', eff_pct), ('trust', trust_pct)]
        dominant = max(pillars, key=lambda p: p[1])[0]
        category_id = CATEGORY_BY_DOMINANT_PILLAR[dominant]

        long_distance = 0 if freq == 'MONTHLY' and difficulty == 'HARD' and 'visit' in title.lower() else 1
        duration = duration_for(freq, difficulty)

        rows_by_freq[freq].append({
            'id': f"snk_{mission_id.lower()}",
            'title': title,
            'description': desc,
            'category_id': category_id,
            'frequency': freq,
            'difficulty': difficulty,
            'xp_reward': xp,
            'comm_pct': comm_pct, 'emo_pct': emo_pct, 'eff_pct': eff_pct, 'trust_pct': trust_pct,
            'verif_type': verif_type,
            'verif_method': verif_raw,
            'verif_level': verif_level,
            'safety_status': safety_status,
            'privacy_rule': privacy_rule,
            'duration': duration,
            'long_distance': long_distance,
        })

    lines = []
    lines.append('-- ============================================================')
    lines.append('-- SNICKYLINK — Migration 0003: Real Snicks Dataset (365 missions)')
    lines.append('-- Auto-generated by scripts/import_snicks.py from SNICKS_updated.xlsx.')
    lines.append('-- Replaces the placeholder 14-Snick seed with the full dataset:')
    lines.append('--   163 Daily (5 XP) / 46 Weekly (15-30 XP) / 30 Monthly (40-100 XP) / 126 Challenge (10-30 XP)')
    lines.append('-- Verification classified per docs/habit-algorithm.md section 1.')
    lines.append('-- ============================================================')
    lines.append('')
    lines.append('PRAGMA foreign_keys = ON;')
    lines.append('')
    lines.append("-- Remove the old placeholder demo Snicks (and their completions/verifications/xp events)")
    lines.append("-- so the couple's journey map is driven entirely by the real dataset going forward.")
    lines.append("DELETE FROM snick_verifications WHERE completion_id IN (SELECT id FROM snick_completions WHERE snick_id LIKE 'snk\\_day%' ESCAPE '\\' OR snick_id LIKE 'snk\\_week%' ESCAPE '\\' OR snick_id LIKE 'snk\\_month%' ESCAPE '\\');")
    lines.append("DELETE FROM couple_xp_events WHERE source_id IN (SELECT id FROM snick_completions WHERE snick_id LIKE 'snk\\_day%' ESCAPE '\\' OR snick_id LIKE 'snk\\_week%' ESCAPE '\\' OR snick_id LIKE 'snk\\_month%' ESCAPE '\\');")
    lines.append("DELETE FROM snick_completions WHERE snick_id LIKE 'snk\\_day%' ESCAPE '\\' OR snick_id LIKE 'snk\\_week%' ESCAPE '\\' OR snick_id LIKE 'snk\\_month%' ESCAPE '\\';")
    lines.append("DELETE FROM snicks WHERE id LIKE 'snk\\_day%' ESCAPE '\\' OR id LIKE 'snk\\_week%' ESCAPE '\\' OR id LIKE 'snk\\_month%' ESCAPE '\\';")
    lines.append('')

    cols = (
        'id, title, description, category_id, frequency, difficulty, xp_reward, '
        'communication_percentage, emotional_connection_percentage, efforts_percentage, trust_percentage, '
        'verification_type, verification_method, verification_level, safety_status, privacy_rule, '
        'duration_minutes, long_distance_supported, location_requirement, sequence_index, map_label, active'
    )

    label_prefix = {'DAILY': 'Day', 'WEEKLY': 'Week', 'MONTHLY': 'Month', 'CHALLENGE': 'Challenge'}

    total = 0
    for freq in ['DAILY', 'WEEKLY', 'MONTHLY', 'CHALLENGE']:
        items = rows_by_freq[freq]
        lines.append(f'-- ---------- {freq} ({len(items)} missions) ----------')
        for i, it in enumerate(items, start=1):
            map_label = f"{label_prefix[freq]} {i}"
            values = [
                sql_str(it['id']), sql_str(it['title']), sql_str(it['description']), sql_str(it['category_id']),
                sql_str(it['frequency']), sql_str(it['difficulty']), it['xp_reward'],
                it['comm_pct'], it['emo_pct'], it['eff_pct'], it['trust_pct'],
                sql_str(it['verif_type']), sql_str(it['verif_method']), sql_str(it['verif_level']),
                sql_str(it['safety_status']), sql_str(it['privacy_rule']),
                it['duration'], it['long_distance'], 'NULL', i, sql_str(map_label), 1,
            ]
            lines.append(f"INSERT INTO snicks ({cols}) VALUES ({', '.join(str(v) for v in values)});")
            total += 1
        lines.append('')

    with open(OUT, 'w') as f:
        f.write('\n'.join(lines) + '\n')

    print(f"Wrote {total} snick rows to {OUT}")
    if unmapped_verification:
        print("WARNING unmapped verification strings:", unmapped_verification)


if __name__ == '__main__':
    main()
